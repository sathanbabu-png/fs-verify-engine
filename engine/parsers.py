"""
Multi-format parsers for financial statement data.
Supports JSON, CSV, and XLSX inputs → FinancialModel.
Uses configurable FieldMapper for robust field name resolution.
"""

import json
import csv
import os
from typing import Dict, Any, Optional, List, Tuple
from .models import (
    FinancialModel, IncomeStatement, BalanceSheet, CashFlowStatement
)
from .field_mapper import (
    FieldMapper, MappingConfig, MappingDiagnostics,
    load_mapping_config, normalize_signs
)


STMT_CLASS_MAP = {
    'income_statement': IncomeStatement,
    'balance_sheet': BalanceSheet,
    'cash_flow': CashFlowStatement,
}

STMT_MODEL_ATTR = {
    'income_statement': 'income_statements',
    'balance_sheet': 'balance_sheets',
    'cash_flow': 'cash_flows',
}


def _parse_number(raw_val: Any) -> float:
    """Parse a number from various formats: $1,234.56, (123), 1,234, -, etc."""
    if raw_val is None:
        return 0.0
    if isinstance(raw_val, (int, float)):
        return float(raw_val)
    s = str(raw_val).strip()
    if not s or s in ('-', '—', '–', 'N/A', 'n/a', '#N/A', '-'):
        return 0.0
    negative = False
    if s.startswith('(') and s.endswith(')'):
        s = s[1:-1]
        negative = True
    s = s.replace('$', '').replace(',', '').replace('%', '').strip()
    if not s:
        return 0.0
    val = float(s)
    return -val if negative else val


def _parse_tabular_data(
    rows: List[List[Any]],
    stmt_type: str,
    mapper: FieldMapper,
    config: MappingConfig,
) -> Tuple[Dict[str, Any], MappingDiagnostics]:
    """
    Parse tabular data (rows = line items, columns = periods).

    Args:
        rows: rows[0] = header row. Column 0 = label, columns 1+ = values.
        stmt_type: 'income_statement', 'balance_sheet', or 'cash_flow'
        mapper: FieldMapper instance
        config: MappingConfig instance

    Returns:
        (dict of period -> statement dataclass, MappingDiagnostics)
    """
    if not rows:
        return {}, MappingDiagnostics(
            statement_type=stmt_type, total_input_fields=0,
            mapped_count=0, unmapped_count=0,
            exact_matches=0, alias_matches=0, fuzzy_matches=0,
        )

    headers = rows[0]
    periods = [str(h).strip() for h in headers[1:] if h and str(h).strip()]

    # Collect all input field names for batch mapping
    input_field_names = []
    for row in rows[1:]:
        if row and row[0] and str(row[0]).strip():
            input_field_names.append(str(row[0]).strip())

    # Map fields
    field_mapping, diagnostics = mapper.map_fields(input_field_names, stmt_type)

    # Build statements
    stmt_class = STMT_CLASS_MAP[stmt_type]
    statements = {}
    for p in periods:
        statements[p] = stmt_class(period=p)

    for row in rows[1:]:
        if not row or not row[0] or not str(row[0]).strip():
            continue
        input_name = str(row[0]).strip()
        internal_field = field_mapping.get(input_name)
        if not internal_field:
            continue

        for j, period in enumerate(periods):
            try:
                val = _parse_number(row[j + 1])
                if hasattr(statements[period], internal_field):
                    setattr(statements[period], internal_field, val)
            except (ValueError, IndexError, TypeError):
                pass

    # Apply sign normalization
    if config.auto_sign_normalization:
        for period in periods:
            stmt = statements[period]
            data = {k: v for k, v in vars(stmt).items() if k != 'period' and isinstance(v, (int, float))}
            normalized = normalize_signs(data, stmt_type, config)
            for k, v in normalized.items():
                setattr(stmt, k, v)

    return statements, diagnostics


# ============================================================================
# JSON Parser
# ============================================================================

def parse_json(
    filepath: str,
    mapping_config: Optional[str] = None,
) -> Tuple[FinancialModel, List[MappingDiagnostics]]:
    """Parse a JSON financial model file."""
    with open(filepath, 'r') as f:
        data = json.load(f)

    config = load_mapping_config(mapping_config)
    mapper = FieldMapper(config)
    all_diagnostics = []

    model = FinancialModel(
        company_name=data.get("company_name", "Unknown"),
        currency=data.get("currency", "USD"),
        unit=data.get("unit", "millions"),
        periods=data.get("periods", []),
        historical_periods=data.get("historical_periods", []),
        projected_periods=data.get("projected_periods", []),
        metadata=data.get("metadata", {}),
    )

    stmt_json_keys = {
        'income_statement': ['income_statements', 'income_statement', 'is', 'pnl', 'p&l'],
        'balance_sheet': ['balance_sheets', 'balance_sheet', 'bs'],
        'cash_flow': ['cash_flows', 'cash_flow', 'cf', 'cash_flow_statement'],
    }

    for stmt_type, json_keys in stmt_json_keys.items():
        stmt_data = None
        for key in json_keys:
            if key in data:
                stmt_data = data[key]
                break
        if not stmt_data:
            continue

        stmt_class = STMT_CLASS_MAP[stmt_type]
        model_attr = STMT_MODEL_ATTR[stmt_type]
        statements = {}

        # Get field names from first period for mapping
        first_period_data = next(iter(stmt_data.values()), {})
        input_fields = [k for k in first_period_data.keys() if k != 'period']
        field_mapping, diagnostics = mapper.map_fields(input_fields, stmt_type)
        all_diagnostics.append(diagnostics)

        for period, items in stmt_data.items():
            stmt = stmt_class(period=period)
            for input_name, value in items.items():
                if input_name == 'period':
                    continue
                # Direct assignment (internal field names)
                if hasattr(stmt, input_name):
                    try:
                        setattr(stmt, input_name, float(value) if value is not None else 0.0)
                    except (ValueError, TypeError):
                        pass
                # Mapped name
                elif input_name in field_mapping:
                    internal = field_mapping[input_name]
                    try:
                        setattr(stmt, internal, float(value) if value is not None else 0.0)
                    except (ValueError, TypeError):
                        pass
            statements[period] = stmt

        setattr(model, model_attr, statements)

    if not model.periods:
        model.periods = model.get_ordered_periods()

    return model, all_diagnostics


def parse_json_string(
    json_str: str,
    mapping_config: Optional[str] = None,
) -> Tuple[FinancialModel, List[MappingDiagnostics]]:
    """Parse a JSON string directly."""
    import tempfile
    data = json.loads(json_str)
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        json.dump(data, f)
        tmppath = f.name
    try:
        return parse_json(tmppath, mapping_config)
    finally:
        try:
            os.unlink(tmppath)
        except (PermissionError, OSError):
            pass


# ============================================================================
# CSV Parser
# ============================================================================

def parse_csv(
    directory: str,
    mapping_config: Optional[str] = None,
) -> Tuple[FinancialModel, List[MappingDiagnostics]]:
    """
    Parse CSV files from a directory.
    Expected: income_statement.csv, balance_sheet.csv, cash_flow.csv
    """
    config = load_mapping_config(mapping_config)
    mapper = FieldMapper(config)
    model = FinancialModel()
    all_diagnostics = []

    file_map = {
        'income_statement': [
            'income_statement.csv', 'income_statements.csv', 'is.csv', 'pnl.csv',
        ],
        'balance_sheet': [
            'balance_sheet.csv', 'balance_sheets.csv', 'bs.csv',
        ],
        'cash_flow': [
            'cash_flow.csv', 'cash_flows.csv', 'cf.csv', 'cash_flow_statement.csv',
        ],
    }

    for stmt_type, filenames in file_map.items():
        filepath = None
        for fn in filenames:
            candidate = os.path.join(directory, fn)
            if os.path.exists(candidate):
                filepath = candidate
                break
        if not filepath:
            continue

        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = [row for row in reader]

        statements, diagnostics = _parse_tabular_data(rows, stmt_type, mapper, config)
        all_diagnostics.append(diagnostics)
        setattr(model, STMT_MODEL_ATTR[stmt_type], statements)

    model.periods = model.get_ordered_periods()
    return model, all_diagnostics


# ============================================================================
# XLSX Parser
# ============================================================================

def parse_xlsx(
    filepath: str,
    mapping_config: Optional[str] = None,
) -> Tuple[FinancialModel, List[MappingDiagnostics]]:
    """
    Parse an Excel file with sheets for each statement.
    Layout: rows = line items, columns = periods.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    config = load_mapping_config(mapping_config)
    mapper = FieldMapper(config)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    model = FinancialModel()
    all_diagnostics = []

    sheet_map = {
        'income_statement': [
            'Income Statement', 'income_statement', 'IS', 'P&L', 'PnL',
            'Income Stmt', 'Profit Loss', 'Profit & Loss',
        ],
        'balance_sheet': [
            'Balance Sheet', 'balance_sheet', 'BS', 'Balance',
        ],
        'cash_flow': [
            'Cash Flow', 'cash_flow', 'CF', 'Cash Flow Statement',
            'Cash Flows', 'SCF', 'Statement of Cash Flows',
        ],
    }

    for stmt_type, sheet_names in sheet_map.items():
        ws = None
        for name in sheet_names:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if not ws:
            continue

        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        statements, diagnostics = _parse_tabular_data(rows, stmt_type, mapper, config)
        all_diagnostics.append(diagnostics)
        setattr(model, STMT_MODEL_ATTR[stmt_type], statements)

    model.periods = model.get_ordered_periods()
    return model, all_diagnostics


# ============================================================================
# Auto Parser
# ============================================================================

def auto_parse(
    filepath: str,
    mapping_config: Optional[str] = None,
) -> Tuple[FinancialModel, List[MappingDiagnostics]]:
    """Auto-detect format and parse with configurable field mapping."""
    if os.path.isdir(filepath):
        return parse_csv(filepath, mapping_config)
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.json':
        return parse_json(filepath, mapping_config)
    elif ext in ('.xlsx', '.xlsm'):
        # Detect if this is a multi-sheet or single-sheet stacked model
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            sheet_names = wb.sheetnames
            sheet_names_lower = [s.lower() for s in sheet_names]
            wb.close()

            # Check if separate statement sheets exist
            has_separate_sheets = any(
                kw in name for name in sheet_names_lower
                for kw in ['income', 'balance', 'cash flow', 'p&l', 'pnl']
                if len(sheet_names) > 1
            )

            if has_separate_sheets and len(sheet_names) > 1:
                return parse_xlsx(filepath, mapping_config)
            else:
                # Single sheet or no recognizable sheet names → try stacked parser
                from .stacked_parser import parse_stacked_sheet
                model, diags = parse_stacked_sheet(filepath, mapping_config=mapping_config)
                # Validate that we got data; if not, fall back to standard parser
                if (model.income_statements or model.balance_sheets or model.cash_flows):
                    return model, diags
                else:
                    return parse_xlsx(filepath, mapping_config)
        except Exception:
            return parse_xlsx(filepath, mapping_config)
    elif ext == '.csv':
        return parse_csv(os.path.dirname(filepath), mapping_config)
    else:
        raise ValueError(f"Unsupported file format: {ext}")

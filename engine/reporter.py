"""
Report generation: Excel with conditional formatting and JSON output.
"""

import json
import os
from typing import Optional
from .engine import VerificationReport
from .models import Severity


def export_json(report: VerificationReport, filepath: str):
    """Export full report as JSON."""
    with open(filepath, 'w') as f:
        f.write(report.to_json())


def export_excel(report: VerificationReport, filepath: str):
    """Export report as Excel with conditional formatting."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Color scheme ──
    FILLS = {
        "pass": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "error": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "critical": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "header": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "section": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    }
    FONTS = {
        "header": Font(bold=True, color="FFFFFF", size=11),
        "section": Font(bold=True, size=11, color="1F4E79"),
        "critical": Font(bold=True, color="FFFFFF"),
        "normal": Font(size=10),
        "title": Font(bold=True, size=14, color="1F4E79"),
    }
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ── Summary Sheet ──
    ws = wb.active
    ws.title = "Summary"
    s = report.summary()

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Verification Report — {s['company_name']}"
    ws['A1'].font = FONTS['title']
    ws['A3'] = "Overall Health:"
    ws['B3'] = s['overall_health']
    health_fill = {"CLEAN": FILLS["pass"], "WARNINGS": FILLS["warning"],
                   "ERRORS_FOUND": FILLS["error"], "CRITICAL": FILLS["critical"]}
    ws['B3'].fill = health_fill.get(s['overall_health'], FILLS["warning"])
    ws['B3'].font = Font(bold=True)

    ws['A4'] = "Timestamp:"
    ws['B4'] = s['timestamp']
    ws['A5'] = "Total Checks:"
    ws['B5'] = s['total_checks']
    ws['A6'] = "Pass Rate:"
    ws['B6'] = f"{s['pass_rate']:.1%}"

    row = 8
    ws[f'A{row}'] = "Severity Breakdown"
    ws[f'A{row}'].font = FONTS['section']
    row += 1
    for sev, count in s['by_severity'].items():
        ws[f'A{row}'] = sev.upper()
        ws[f'B{row}'] = count
        if sev in FILLS:
            ws[f'A{row}'].fill = FILLS[sev]
        row += 1

    row += 1
    ws[f'A{row}'] = "Category Breakdown"
    ws[f'A{row}'].font = FONTS['section']
    row += 1
    for cat, stats in s['by_category'].items():
        ws[f'A{row}'] = cat
        ws[f'B{row}'] = f"{stats['passed']}/{stats['total']} ({stats['pass_rate']:.0%})"
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30

    # ── Detail Sheet ──
    ws2 = wb.create_sheet("Check Results")
    headers = ["Check ID", "Check Name", "Category", "Period", "Severity",
               "Message", "Expected", "Actual", "Delta", "Delta %"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = FONTS['header']
        cell.fill = FILLS['header']
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, r in enumerate(report.results, 2):
        values = [
            r.check_id, r.check_name, r.category.value, r.period,
            r.severity.value.upper(), r.message,
            r.expected_value, r.actual_value, r.delta,
            f"{r.delta_pct:.4%}" if r.delta_pct is not None else None,
        ]
        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = FONTS['normal']
            cell.border = thin_border
        # Apply row fill based on severity
        sev_key = r.severity.value
        if sev_key in FILLS:
            for col in range(1, len(headers) + 1):
                ws2.cell(row=i, column=col).fill = FILLS[sev_key]
                if sev_key == "critical":
                    ws2.cell(row=i, column=col).font = FONTS['critical']

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws2.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(len(report.results) + 2, 100))
        )
        ws2.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── Failures Only Sheet ──
    ws3 = wb.create_sheet("Failures")
    failures = report.get_failures()
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = FONTS['header']
        cell.fill = FILLS['header']
        cell.border = thin_border

    for i, r in enumerate(failures, 2):
        values = [
            r.check_id, r.check_name, r.category.value, r.period,
            r.severity.value.upper(), r.message,
            r.expected_value, r.actual_value, r.delta,
            f"{r.delta_pct:.4%}" if r.delta_pct is not None else None,
        ]
        for col, val in enumerate(values, 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.border = thin_border
            sev_key = r.severity.value
            if sev_key in FILLS:
                cell.fill = FILLS[sev_key]

    wb.save(filepath)
